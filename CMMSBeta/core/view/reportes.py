import datetime
import string
import random
import io
import base64

from django.shortcuts import redirect, render
from django.http import HttpResponse, FileResponse, HttpResponseBadRequest
from django.template.loader import get_template
from django.contrib.auth.decorators import login_required
from django.db.models import Count
from django.core.exceptions import ObjectDoesNotExist

from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, Alignment, Border, PatternFill
from openpyxl.styles.borders import Border, Side

from programacion_labor.models import DetalleLabor, Programacion, TipoLabor
from tractor.models import ReporteTractor

import plotly.graph_objs as go
import matplotlib.pyplot as plt
import numpy as np

from xhtml2pdf import pisa

# Reporte En Excel
@login_required(login_url='login', redirect_field_name='home')
def exportar(request, fecha_inicio=None, fecha_fin=None):
    if request.method == 'POST':
        try:
            fecha_inicio = request.POST.get('fecha_inicio', None)
            fecha_fin = request.POST.get('fecha_fin', None)
            usuario_id = request.POST.get('usuario', None)
            labor_id = request.POST.get('labor', None)
            horas_uso = request.POST.get('horas_tractor', None)

            # Generar el sufijo para el nombre del archivo
            sufijo = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
            nombre_archivo = f"reporte-{sufijo}.xlsx"

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

            nuevoLibro = Workbook()
            hojaActiva = nuevoLibro.active

            # Establecer estilos y formatos
            bold_font = Font(bold=True)
            estilo_tabla = NamedStyle(name="estilo_tabla")
            estilo_tabla.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            estilo_tabla.border = Border(left=Side(border_style='thin', color='000000'),
                                         right=Side(border_style='thin', color='000000'),
                                         top=Side(border_style='thin', color='000000'),
                                         bottom=Side(border_style='thin', color='000000'))

            nuevoLibro.add_named_style(estilo_tabla)

            # Escribir encabezado
            encabezado = ['Sede', 'Implemento', 'Horometro Inicial', 'Horometro Final', 'Fundo', 'Tipo de labor',
                          'Lote', 'Tractor', 'Usuario', 'Tractorista', 'Solicitante', 'Fecha', 'Turno']
            hojaActiva.append(encabezado)

            # Establecer estilo y formato para el encabezado
            for celda in hojaActiva[1]:
                celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                celda.font = bold_font
                celda.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                celda.border = Border(left=Side(border_style='thin', color='000000'),
                                      right=Side(border_style='thin', color='000000'),
                                      top=Side(border_style='thin', color='000000'),
                                      bottom=Side(border_style='thin', color='000000'))

            # Obtener datos de la base de datos
            detalles_labor_query = DetalleLabor.objects.all()

            # Filtrar detalles labor según los parámetros recibidos
            if fecha_inicio:
                detalles_labor_query = detalles_labor_query.filter(idprogramacion__fechahora__gte=fecha_inicio)
            if fecha_fin:
                detalles_labor_query = detalles_labor_query.filter(idprogramacion__fechahora__lte=fecha_fin)
            if usuario_id:
                detalles_labor_query = detalles_labor_query.filter(idprogramacion__idusuario=usuario_id)
            if labor_id:
                detalles_labor_query = detalles_labor_query.filter(idprogramacion__idtipolabor=labor_id)
            if horas_uso:
                detalles_labor_query = detalles_labor_query.filter(idprogramacion__idtractor__horauso__gte=horas_uso)

            detalles_labor = detalles_labor_query.values_list('idimplemento__implemento',
                                                               'idprogramacion').order_by('-idprogramacion')

            # Verificar si se encontraron datos después de aplicar los filtros
            if not detalles_labor.exists():
                return HttpResponse("No se encontraron datos para los filtros aplicados.", status=404)

            for detalle in detalles_labor:
                detalle_labor = list(detalle)
                implemento_nombre = detalle_labor[0]
                id_programacion = detalle_labor[1]

                try:
                    programacion = Programacion.objects.get(idprogramacion=id_programacion)
                    tipolabor = TipoLabor.objects.get(idtipolabor=programacion.idtipolabor_id)

                    solicitante_nombre = ""
                    if programacion.idsolicitante:
                        solicitante_nombre = f"{programacion.idsolicitante.idpersona.nombres} {programacion.idsolicitante.idpersona.apellidos}"

                    tractorista_nombre = ""
                    if programacion.idtractorista:
                        tractorista_nombre = f"{programacion.idtractorista.idpersona.nombres} {programacion.idtractorista.idpersona.apellidos}"

                    reporteTractores = ReporteTractor.objects.filter(idprogramacion=id_programacion)

                    for reporteTractor in reporteTractores:
                        fecha_formateada = str(programacion.fechahora)
                        estado = 'Activo' if detalle_labor[-1] else 'Inactivo'
                        detalle_labor[-1] = estado

                        horometro_inicial = ""
                        horometro_final = ""
                        if reporteTractor.horometroinicial:
                            horometro_inicial = reporteTractor.horometroinicial
                        if reporteTractor.horometrofinal:
                            horometro_final = reporteTractor.horometrofinal

                        datos_programacion = [
                            programacion.idlote.idfundo.idsede.sede,
                            horometro_inicial,
                            horometro_final,
                            programacion.idlote.idfundo.fundo,
                            tipolabor.tipolabor,
                            programacion.idlote.lote if programacion.idlote else '',
                            programacion.idtractor.nrotractor if programacion.idtractor else '',
                            programacion.idusuario.username if programacion.idusuario else '',
                            tractorista_nombre,
                            solicitante_nombre,
                            fecha_formateada,
                            programacion.turno,
                        ]
                        hojaActiva.append([datos_programacion[0]] + [detalle_labor[0]] + datos_programacion[1:])

                except Programacion.DoesNotExist:
                    print("No se encontró la programación con el ID proporcionado.")
                except TipoLabor.DoesNotExist:
                    print("No se encontró el tipo de labor con el ID proporcionado.")
                except ReporteTractor.DoesNotExist:
                    print("No se encontró el reporte del tractor asociado.")

            # Ajustar estilo y ancho de columnas
            for row in hojaActiva.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)

            for col in hojaActiva.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                hojaActiva.column_dimensions[column].width = adjusted_width

            for fila in hojaActiva.iter_rows(min_row=2, max_row=hojaActiva.max_row, min_col=1,
                                              max_col=hojaActiva.max_column):
                for celda in fila:
                    celda.style = "estilo_tabla"

            nuevoLibro.save(response)
            return response

        except Exception as e:
            print(f"Se produjo un error: {str(e)}")
            return HttpResponse("Se produjo un error al exportar los datos.")


# Reporte en PDF
def reportePDF(request):
    # Obtener parámetros del formulario
    fecha_inicio = request.POST.get('fecha_inicio', None)
    fecha_fin = request.POST.get('fecha_fin', None)
    usuario_id = request.POST.get('usuario', None)
    labor_id = request.POST.get('labor', None)
    horas_uso = request.POST.get('horas_tractor', None)

    # Validar y convertir fechas
    fecha_inicial = None
    fecha_final = None
    if fecha_inicio:
        try:
            fecha_inicial = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        except ValueError:
            return HttpResponse("La fecha de inicio proporcionada no es válida.", status=400)

    if fecha_fin:
        try:
            fecha_final = datetime.strptime(fecha_fin, '%Y-%m-%d')
        except ValueError:
            return HttpResponse("La fecha de fin proporcionada no es válida.", status=400)

    # Filtrar detalles de labor según los parámetros
    detalles_labor = DetalleLabor.objects.all().order_by('-idprogramacion')

    if fecha_inicial and fecha_final:
        detalles_labor = detalles_labor.filter(idprogramacion__fechahora__range=(fecha_inicial, fecha_final))
        if not detalles_labor.exists():
            return HttpResponse("No hay datos disponibles para las fechas especificadas.", status=400)

    if usuario_id:
        detalles_labor = detalles_labor.filter(idprogramacion__idusuario=usuario_id)
        if not detalles_labor.exists():
            return HttpResponse("No hay datos disponibles para el usuario seleccionado.", status=400)

    if labor_id:
        detalles_labor = detalles_labor.filter(idprogramacion__idtipolabor_id=labor_id)
        if not detalles_labor.exists():
            return HttpResponse("No hay datos disponibles para el tipo de labor seleccionado.", status=400)

    if horas_uso:
        detalles_labor = detalles_labor.filter(idprogramacion__idtractor__horauso__gte=horas_uso)
        if not detalles_labor.exists():
            return HttpResponse("No hay datos disponibles para las horas de uso seleccionadas.", status=400)

    # Generar el PDF
    template = get_template('reportes/reporte_pdf.html')
    context = {'detalles_labor': detalles_labor}
    html = template.render(context)

    buffer = io.BytesIO()
    pisa.CreatePDF(html, dest=buffer, encoding='utf-8')
    buffer.seek(0)

    # Devolver el PDF como respuesta
    return FileResponse(buffer, as_attachment=True, filename="reporte.pdf")


# Reporte Gráfico
from django.http import JsonResponse

def reporteGrafico(request):
    fecha = request.POST.get('fecha_grafico')
    if fecha is None:
        fecha = datetime.date.today()
        print (fecha)

    if fecha:
        solicitudes = Programacion.objects.filter(fechahora=fecha)

        if not solicitudes.exists():
            return JsonResponse({"error": "No hay solicitudes en la fecha solicitada"}, status=400)

        nombres_solicitudes = solicitudes.values('idsolicitante__idpersona__nombres') \
            .annotate(num_solicitudes=Count('idsolicitante')) \
            .order_by()

        data = [{"nombre": nombre['idsolicitante__idpersona__nombres'], "num_solicitudes": nombre['num_solicitudes']} for nombre in nombres_solicitudes]

        # Redirigir a una nueva página con los datos del gráfico
        return redirect('home_with_datagrafic', datagrafic=data)

    else:
        return JsonResponse({"error": "No se proporcionó una fecha válida"}, status=400)

