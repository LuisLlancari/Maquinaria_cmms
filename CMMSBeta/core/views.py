from django.http import HttpResponse
from openpyxl import Workbook
from django.shortcuts import render
from openpyxl.styles import Font, NamedStyle, Alignment, Border, Side, PatternFill
from openpyxl.styles.borders import Border, Side
from programacion_labor.models import DetalleLabor, Programacion, TipoLabor
from django.contrib.auth.decorators import login_required
from datetime import datetime
from django.template.loader import get_template
from django.http import FileResponse
from xhtml2pdf import pisa
import io



#IMPORTACIONES EXTERNAS
import string
import random

@login_required(login_url='login', redirect_field_name='home')
def home(request):
  usuario = request.user
  print("Usuario logueado:", usuario)
  return render(request, 'core/home.html', {'user': usuario})

@login_required(login_url='login', redirect_field_name='home')
def test(request):
    return render(request, 'core/test.html')

@login_required(login_url='login', redirect_field_name='home')
def exportar(request, fecha_inicio=None, fecha_fin=None):
    if request.method == 'POST':
        try:  
            sufijo = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
            nombre_archivo = f"reporte-{sufijo}.xlsx"
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

            nuevoLibro = Workbook()
            hojaActiva = nuevoLibro.active

            bold_font = Font(bold=True)

            # Estilo de tabla
            estilo_tabla = NamedStyle(name="estilo_tabla")
            estilo_tabla.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            estilo_tabla.border = Border(left=Side(border_style='thin', color='000000'),
                                        right=Side(border_style='thin', color='000000'),
                                        top=Side(border_style='thin', color='000000'),
                                        bottom=Side(border_style='thin', color='000000'))

            nuevoLibro.add_named_style(estilo_tabla)

            # Escribir encabezado
            encabezado = ['Nro Detalle Labor', 'Implemento', 'Programacion', 'Tipo de labor', 'Lote', 'Tractor', 'Usuario', 'Tractorista', 'Solicitante', 'Fecha', 'Turno', 'Horas Uso', 'Estado']
            hojaActiva.append(encabezado)
            # Establecer estilo y formato para el encabezado
            for celda in hojaActiva[1]:
                celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                celda.font = bold_font
                celda.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                celda.border = Border(left=Side(border_style='thin', color='000000'),
                                    right=Side(border_style='thin', color='000000'),
                                    top=Side(border_style='thin', color='000000'),
                                    bottom=Side(border_style='thin', color='000000'))


            # Obtener datos de la base de datos
            detalles_labor_query = DetalleLabor.objects.all()

            detalles_labor = detalles_labor_query.values_list('iddetlabor', 'idimplemento__implemento', 'idprogramacion', 'horadeuso', 'estado').order_by('-idprogramacion')

            # Agregar datos a la hoja
            for detalle in detalles_labor:
                detalle_labor = list(detalle)
                implemento_nombre = detalle_labor[1]  # Nombre del implemento
                id_programacion = detalle_labor[2]
                programacion = Programacion.objects.get(idprogramacion=id_programacion)
                tipolabor = TipoLabor.objects.get(idtipolabor=programacion.idtipolabor_id)  
                solicitante_nombre = f"{programacion.idsolicitante.nombres} {programacion.idsolicitante.apellidos}" if programacion.idsolicitante else ''  
                tractorista_nombre = f"{programacion.idtractorista.nombres} {programacion.idtractorista.apellidos}" if programacion.idtractorista else ''  

                fecha_formateada = str(programacion.fechahora)
                estado = 'Activo' if detalle_labor[-1] else 'Inactivo'

                detalle_labor[-1] = estado

                datos_programacion = [
                    id_programacion,  # ID de la programación
                    tipolabor.tipolabor,  
                    programacion.idlote.lote if programacion.idlote else '',
                    programacion.idtractor.nrotractor if programacion.idtractor else '',
                    programacion.idusuario.username if programacion.idusuario else '',
                    tractorista_nombre,  
                    solicitante_nombre,  
                    fecha_formateada,  
                    programacion.turno,
                ]
                hojaActiva.append([detalle_labor[0]] + [implemento_nombre] + datos_programacion + detalle_labor[-2:])

            
            for row in hojaActiva.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True)

            for col in hojaActiva.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                hojaActiva.column_dimensions[column].width = adjusted_width

        
            for fila in hojaActiva.iter_rows(min_row=2, max_row=hojaActiva.max_row, min_col=1, max_col=hojaActiva.max_column):
                for celda in fila:
                    celda.style = "estilo_tabla"

            nuevoLibro.save(response)
            return response

        except Exception as e:

            print(f"Se produjo un error: {str(e)}")

            return HttpResponse("Se produjo un error al exportar los datos.")


def reportePDF(request):
# Obtener datos de la base de datos
    detalles_labor = DetalleLabor.objects.all().order_by('-idprogramacion')

    # Renderizar el template con los datos
    template = get_template('core/reporte_pdf.html')
    context = {'detalles_labor': detalles_labor}
    html = template.render(context)

    # Crear un archivo PDF en memoria
    buffer = io.BytesIO()
    pisa.CreatePDF(html, dest=buffer, encoding='utf-8')

    # Devolver el PDF como respuesta
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename="reporte.pdf")

